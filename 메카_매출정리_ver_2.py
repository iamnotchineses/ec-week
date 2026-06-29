import html
import io
import re
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

st.set_page_config(page_title="메카 매출 대시보드", page_icon="📊", layout="wide")

APP_DIR = Path(__file__).parent




# -----------------------------
# Style
# -----------------------------
st.markdown(
    """
    <style>
    .main .block-container {padding-top: 1.5rem; padding-bottom: 2rem;}
    div[data-testid="stMetric"] {
        background: #ffffff;
        border: 1px solid #e8eef5;
        padding: 14px 14px;
        border-radius: 16px;
        box-shadow: 0 4px 16px rgba(15, 23, 42, 0.04);
    }
    div[data-testid="stMetric"] label {color:#64748b; font-size:0.85rem;}
    div[data-testid="stMetricValue"] {
        font-size: 1.3rem;
        font-weight: 700;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }
    div[data-testid="stMetricValue"] > div {
        overflow: hidden;
        text-overflow: ellipsis;
        white-space: nowrap;
    }
    .section-title {
        font-size: 1.15rem;
        font-weight: 800;
        margin: 1.2rem 0 .4rem 0;
        color: #0f172a;
    }
    .hint {
        color: #64748b;
        font-size: .9rem;
    }
    .printtbl { border-collapse: collapse; width: 100%; font-size: 12px; margin: 6px 0; }
    .printtbl th, .printtbl td { border: 1px solid #d0d7de; padding: 4px 8px; text-align: right; white-space: nowrap; }
    .printtbl th { background: #f1f5f9; text-align: center; font-weight: 700; }
    .printtbl td.l { text-align: left; }
    .printtbl tbody tr:nth-child(even) { background: #fafbfc; }
    @media print {
        @page { size: A4 landscape; }
        [data-testid="stSidebar"], [data-testid="stToolbar"], [data-testid="stHeader"],
        [data-testid="stDecoration"], header, footer, [data-testid="stStatusWidget"] { display: none !important; }
        .stApp, .main, .block-container { background: #fff !important; padding-top: 0 !important; }
        section.main div.block-container, .main .block-container, .block-container { max-width: 100% !important; }
        * { -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }
        /* 차트 폭만 용지에 맞춤 (레이아웃은 화면 그대로 둠) */
        [data-testid="stPlotlyChart"], [data-testid="stPlotlyChart"] > div,
        .js-plotly-plot, .js-plotly-plot .main-svg, .svg-container, .plot-container {
            width: 100% !important; max-width: 100% !important;
        }
        /* 차트/표 한 줄만 잘리지 않게. 큰 블록은 자연스럽게 흐르게(빈 여백 방지) */
        [data-testid="stPlotlyChart"], .js-plotly-plot { page-break-inside: avoid; }
        .printtbl { page-break-inside: auto; }
        .printtbl thead { display: table-header-group; }   /* 페이지 넘어가도 헤더 반복 */
        .printtbl tr { page-break-inside: avoid; }          /* 행 중간 안 잘림 */
        .section-title { page-break-after: avoid; break-after: avoid; }  /* 제목 혼자 페이지 끝에 안 남게 */
        h3, h4, .stMarkdown p strong { page-break-after: avoid; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# -----------------------------
# Helpers
# -----------------------------
def clean_col_name(col: str) -> str:
    return re.sub(r"\s+", " ", str(col).replace("\n", " ")).strip()


def _xlsx_rows_fast(data: bytes) -> list:
    """openpyxl read_only + data_only 로 첫 시트를 값만 빠르게 읽음.
    임베드 이미지/수식이 많은 무거운 .xlsx 도 빠르고 메모리 적게 읽는다."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    while rows and all(c is None for c in rows[-1]):  # 빈 꼬리행 제거
        rows.pop()
    return rows


def _html_rows_fast(text: str) -> list:
    """HTML 위장 .xls 의 <table> 를 lxml 스트리밍 파싱(<tr> 단위로 읽고 해제 → 100MB+도 메모리 적게)."""
    from lxml import etree
    rows = []
    ctx = etree.iterparse(io.BytesIO(text.encode("utf-8")), events=("end",),
                          tag="tr", html=True, recover=True, encoding="utf-8")
    for _, tr in ctx:
        cells = []
        for cell in tr:
            if isinstance(cell.tag, str) and cell.tag in ("td", "th"):
                txt = "".join(cell.itertext()).strip()
                cells.append(txt if txt != "" else None)
        if cells:
            rows.append(cells)
        tr.clear()
        while tr.getprevious() is not None:  # 처리한 행 메모리 해제
            del tr.getparent()[0]
    return rows


def _dedupe_cols(cols):
    seen, out = {}, []
    for c in cols:
        c = c if (c is not None and str(c).strip() != "") else "col"
        if c in seen:
            seen[c] += 1; out.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0; out.append(c)
    return out


def read_excel_smart(file_obj) -> pd.DataFrame:
    """Read Excel where the first row may be blank and the actual header starts later.
    openpyxl read_only(값만)로 빠르게 읽고, 실패 시 pandas 로 폴백."""
    data = file_obj.read() if hasattr(file_obj, "read") else file_obj
    try:
        rows = _xlsx_rows_fast(data)
    except Exception:
        raw = pd.read_excel(io.BytesIO(data), sheet_name=0, header=None)
        rows = [list(r) for r in raw.where(pd.notna(raw), None).values.tolist()]
    if not rows:
        return pd.DataFrame()
    keywords = {"주문번호", "쇼핑몰", "브랜드", "수량", "최종판매가", "출고날짜"}
    header_row = 0
    for i in range(min(10, len(rows))):
        values = set(str(x).strip() for x in rows[i] if x is not None)
        if len(values & keywords) >= 3:
            header_row = i
            break
    cols = _dedupe_cols([str(c).strip() if c is not None else "" for c in rows[header_row]])
    df = pd.DataFrame(rows[header_row + 1:], columns=cols, dtype=object)
    df.columns = [clean_col_name(c) for c in df.columns]
    # 엑셀 AA열(27번째 = 0-based 26)을 '위치 그대로' 확보(헤더 비면 Unnamed 로 잡혀 삭제되므로)
    aa_series = df.iloc[:, 26].copy() if df.shape[1] > 26 else None
    df = df.loc[:, ~pd.Series(df.columns).astype(str).str.startswith("Unnamed").values]
    df = df.dropna(how="all")
    if aa_series is not None and "대분류" not in df.columns:
        df["대분류"] = aa_series.reindex(df.index)
    return df


def find_col(df: pd.DataFrame, candidates: list[str], fallback_contains: str | None = None) -> str | None:
    cols = list(df.columns)
    for c in candidates:
        if c in cols:
            return c
    if fallback_contains:
        for c in cols:
            if fallback_contains in c:
                return c
    return None


def to_number(s: pd.Series) -> pd.Series:
    if s is None:
        return pd.Series(dtype=float)
    return pd.to_numeric(
        s.astype(str).str.replace(",", "", regex=False).str.replace("%", "", regex=False),
        errors="coerce",
    ).fillna(0)


def money(v) -> str:
    try:
        return f"{float(v):,.0f}원"
    except Exception:
        return "0원"


def num(v) -> str:
    try:
        return f"{float(v):,.0f}"
    except Exception:
        return "0"


def eok(v) -> str:
    """금액 표기: 1억 이상이면 'X.X억', 그 미만이면 전체 숫자(콤마)."""
    try:
        v = float(v)
    except Exception:
        return "0"
    if pd.isna(v):
        return "-"
    if abs(v) >= 1e8:
        return f"{v / 1e8:.1f}억"
    return f"{v:,.0f}"


def to_line(model) -> str:
    """모델명에서 끝의 사이즈 '(...)' 를 떼어 라인명으로 변환.
    예: 'COHBU M26388 ALI BLANC/BLEU CIEL (XL)' -> 'COHBU M26388 ALI BLANC/BLEU CIEL'
        'K100979-001 (44)' -> 'K100979-001'
    """
    return re.sub(r"\s*\([^()]*\)\s*$", "", str(model)).strip()


line_map: dict = {}  # 모델명(정규화) → 라인명. 로드 시 (이미지 3번째 시트 + 재고)로 채움.


def _norm_model(s) -> str:
    return re.sub(r"\s+", " ", str(s)).strip()


def _line_of(model) -> str:
    """모델명 → 라인명. line_map(이미지 3번째 시트/재고) 우선, 없으면 끝의 사이즈 '(...)' 제거."""
    nm = _norm_model(model)
    if nm in line_map:
        return line_map[nm]
    nm2 = _norm_model(to_line(model))
    if nm2 in line_map:
        return line_map[nm2]
    return to_line(model)


def pct(v) -> str:
    try:
        if pd.isna(v) or np.isinf(v):
            return "-"
        return f"{float(v):,.1f}%"
    except Exception:
        return "-"


def growth_pct(v) -> str:
    """Format growth-rate columns like ▲ 18.0% / ▼ 15.0% with one decimal."""
    try:
        if pd.isna(v) or np.isinf(v):
            return "-"
        value = float(v)
        if value > 0:
            return f"▲ {abs(value):,.1f}%"
        if value < 0:
            return f"▼ {abs(value):,.1f}%"
        return "0.0%"
    except Exception:
        return "-"


def add_rate(df: pd.DataFrame, current_col: str, prev_col: str, out_col: str = "YoY 신장률") -> pd.DataFrame:
    prev = df[prev_col].replace(0, np.nan)
    df[out_col] = ((df[current_col] - df[prev_col]) / prev.abs()) * 100
    return df


def sort_desc(df: pd.DataFrame, by: str) -> pd.DataFrame:
    if by in df.columns:
        return df.sort_values(by=by, ascending=False, na_position="last")
    return df


# 표/차트의 '총매출' 라벨 접두어 — 본문에서 모드(주간/월간)에 따라 재설정됨
INTERVAL_LABEL = "주간"


_MONEY_KW = ["판매가", "매출", "수익원", "원가", "증감", "객단가", "수수료액", "배송비", "신장액", "총매출", "총원가", "금액", "목표", "실제"]


def _is_money_col(c) -> bool:
    """금액(억 단위 표시) 컬럼인지. %(률/율/비중)는 제외."""
    c = str(c)
    if any(k in c for k in ("률", "율", "비중", "Rate", "달성")):
        return False
    return (any(k in c for k in _MONEY_KW)
            or bool(re.fullmatch(r"\d{4}년", c))
            or bool(re.fullmatch(r"\d{1,2}월\s?\d{1,2}주차", c))
            or bool(re.fullmatch(r"\d{4}년\s?\d{1,2}월", c)))


PRINT_MODE = False  # 사이드바에서 켜면 표를 정적 HTML(인쇄/PDF용)로 렌더


def _fmt_cell(col, val) -> str:
    """인쇄용 HTML 표의 셀 표시값(숫자 콤마/%/부호)."""
    if pd.isna(val):
        return "-"
    cs = str(col)
    try:
        if cs in ("Rank", "순위"):
            return f"{int(val)}"
        if "신장률" in cs or "신장율" in cs or "대비" in cs:
            return f"{val:+.1f}%"
        if any(k in cs for k in ["률", "율", "비중", "Rate", "달성"]):
            return f"{val:.1f}%"
        if cs in ("수량", "주문수", "라인수"):
            return f"{val:,.0f}"
        if _is_money_col(cs):
            return f"{val:,.0f}"
    except Exception:
        pass
    return str(val)


def _df_to_html(df: pd.DataFrame) -> str:
    """데이터프레임 → 인쇄 친화 정적 HTML 표."""
    import html as _h
    name_cols = {"쇼핑몰", "브랜드", "대분류", "모델명", "요일", "공식/병행", "라인명"}
    head = "".join(f"<th>{_h.escape(str(c))}</th>" for c in df.columns)
    body = []
    for _, r in df.iterrows():
        tds = []
        for c in df.columns:
            cls = " class='l'" if str(c) in name_cols else ""
            tds.append(f"<td{cls}>{_h.escape(_fmt_cell(c, r[c]))}</td>")
        body.append("<tr>" + "".join(tds) + "</tr>")
    return (f"<table class='printtbl'><thead><tr>{head}</tr></thead>"
            f"<tbody>{''.join(body)}</tbody></table>")


def format_table(df: pd.DataFrame) -> pd.DataFrame:
    """표시용 데이터프레임: 값은 '숫자형'으로 유지해 st.dataframe 정렬이 숫자로 되게 한다.
    (콤마/억/% 표시는 _table_config 의 column_config 가 담당) 금액은 '억 단위'로 환산해 둔다."""
    out = df.copy()
    # 수량을 최종판매가(총매출) 바로 왼쪽으로 이동 (모든 표 일괄)
    if "수량" in out.columns and "최종판매가" in out.columns:
        cols = list(out.columns)
        cols.remove("수량")
        cols.insert(cols.index("최종판매가"), "수량")
        out = out[cols]
    rename_map = {}
    if "최종판매가" in out.columns:
        rename_map["최종판매가"] = f"{INTERVAL_LABEL} 총매출"
    if "수익원(실배송비)" in out.columns:
        rename_map["수익원(실배송비)"] = "수익원"
    if rename_map:
        out = out.rename(columns=rename_map)
    pct_keywords = ["률", "율", "비중", "Rate"]
    for c in out.columns:
        cs = str(c)
        if cs in ("Rank", "순위"):
            out[c] = pd.to_numeric(out[c], errors="coerce").round(0)
        elif "신장률" in cs or "신장율" in cs or "대비" in cs or any(k in cs for k in pct_keywords):
            out[c] = pd.to_numeric(out[c], errors="coerce").replace([np.inf, -np.inf], np.nan).round(1)
        elif _is_money_col(cs):
            out[c] = pd.to_numeric(out[c], errors="coerce").replace([np.inf, -np.inf], np.nan).round(0)
        elif cs in ("수량", "주문수", "라인수"):
            out[c] = pd.to_numeric(out[c], errors="coerce").round(0)
        # else: 문자열 컬럼은 그대로
    return out


def _table_config(df: pd.DataFrame) -> dict:
    """컬럼명 패턴으로 st.column_config 생성 (정렬은 숫자, 표시는 억/콤마/%)."""
    cfg = {}
    for c in df.columns:
        cs = str(c)
        if cs in ("Rank", "순위"):
            cfg[c] = st.column_config.NumberColumn("#" if cs == "Rank" else cs, format="%d")
        elif "신장률" in cs or "신장율" in cs or "대비" in cs:
            cfg[c] = st.column_config.NumberColumn(format="%+.1f%%")
        elif any(k in cs for k in ["률", "율", "비중", "Rate", "달성"]):
            cfg[c] = st.column_config.NumberColumn(format="%.1f%%")
        elif cs in ("수량", "주문수", "라인수"):
            cfg[c] = st.column_config.NumberColumn(format="%,.0f")
        elif _is_money_col(cs):
            cfg[c] = st.column_config.NumberColumn(format="%,.0f")
    return cfg


def show_table(df: pd.DataFrame, **kwargs) -> None:
    """format_table + 자동 column_config 로 정렬 가능한 표를 렌더 (인쇄모드면 정적 HTML)."""
    d = format_table(df)
    if PRINT_MODE:
        st.markdown(_df_to_html(d), unsafe_allow_html=True)
        return
    kwargs.setdefault("hide_index", True)
    kwargs.setdefault("use_container_width", True)
    cfg = _table_config(d)
    if "column_config" in kwargs and kwargs["column_config"]:
        cfg = {**cfg, **kwargs.pop("column_config")}
    else:
        kwargs.pop("column_config", None)
    st.dataframe(d, column_config=cfg, **kwargs)


def aggregate(df: pd.DataFrame, group_cols: list[str], metric_cols: dict) -> pd.DataFrame:
    agg_spec = {}
    for out, col in metric_cols.items():
        if col and col in df.columns:
            agg_spec[out] = (col, "sum")
    result = df.groupby(group_cols, dropna=False).agg(**agg_spec).reset_index()
    if "최종판매가" in result.columns and "수량" in result.columns:
        result["객단가"] = np.where(result["수량"] != 0, result["최종판매가"] / result["수량"], 0)
    if "수익원(실배송비)" in result.columns and "최종판매가" in result.columns:
        result["수익률"] = np.where(result["최종판매가"] != 0, result["수익원(실배송비)"] / result["최종판매가"] * 100, 0)
    total = result["최종판매가"].sum() if "최종판매가" in result.columns else 0
    if total != 0 and "최종판매가" in result.columns:
        result["매출비중"] = result["최종판매가"] / total * 100
    return sort_desc(result, "최종판매가")



def rank_table(df: pd.DataFrame, name_col: str) -> pd.DataFrame:
    """Add rank and replace the displayed name with rank order prefix."""
    out = df.copy().reset_index(drop=True)
    out.insert(0, "Rank", np.arange(1, len(out) + 1))
    if name_col in out.columns:
        clean_name = out[name_col].astype(str).str.replace(r"^\s*\d+\s*[\.\)\-_/]*\s*", "", regex=True)
        out[name_col] = out["Rank"].astype(str) + ". " + clean_name
    return out


def top_sales_table(df: pd.DataFrame, group_cols: list[str], topn: int = 30, sort_by: str = "최종판매가") -> pd.DataFrame:
    table = aggregate(df, group_cols, metric_cols)
    if sort_by in table.columns:
        table = table.sort_values(sort_by, ascending=False, na_position="last")
    table = table.head(topn).reset_index(drop=True)
    table.insert(0, "Rank", np.arange(1, len(table) + 1))
    return table

def yoy_by_group(df: pd.DataFrame, group_col: str, base_year: int, metric_col: str) -> pd.DataFrame:
    prev_year = base_year - 1
    temp = df[df["연도"].isin([prev_year, base_year])]
    pivot = temp.pivot_table(index=group_col, columns="연도", values=metric_col, aggfunc="sum", fill_value=0).reset_index()
    if prev_year not in pivot.columns:
        pivot[prev_year] = 0
    if base_year not in pivot.columns:
        pivot[base_year] = 0
    pivot = pivot.rename(columns={prev_year: f"{prev_year}년", base_year: f"{base_year}년"})
    pivot["YoY 신장액"] = pivot[f"{base_year}년"] - pivot[f"{prev_year}년"]
    pivot = add_rate(pivot, f"{base_year}년", f"{prev_year}년")
    return sort_desc(pivot, f"{base_year}년")


def trend_by_group(df: pd.DataFrame, group_col: str, metric_col: str, topn=None) -> pd.DataFrame:
    """그룹별 연도 추이: 데이터에 존재하는 모든 연도를 열로 펼친다. 최신 연도 기준 내림차순."""
    years = sorted(int(y) for y in df["연도"].dropna().unique())
    pivot = df.pivot_table(
        index=group_col, columns="연도", values=metric_col, aggfunc="sum", fill_value=0
    ).reset_index()
    for y in years:
        if y not in pivot.columns:
            pivot[y] = 0
    pivot = pivot.rename(columns={y: f"{y}년" for y in years})
    year_cols = [f"{y}년" for y in years]
    pivot = pivot[[group_col] + year_cols]
    if year_cols:
        pivot = pivot.sort_values(year_cols[-1], ascending=False, na_position="last")
    pivot = pivot.reset_index(drop=True)
    if topn:
        pivot = pivot.head(topn)
    return pivot


def wow_by_group(df: pd.DataFrame, group_col: str, metric_col: str, week_order: list, topn=None) -> pd.DataFrame:
    """그룹별 주차 추이: week_order(시간순 주차 라벨)대로 열을 펼친다. 최신 주차 기준 내림차순."""
    pivot = df.pivot_table(
        index=group_col, columns="주차", values=metric_col, aggfunc="sum", fill_value=0
    ).reset_index()
    week_cols = [w for w in week_order if w in pivot.columns]
    pivot = pivot[[group_col] + week_cols]
    if week_cols:
        pivot = pivot.sort_values(week_cols[-1], ascending=False, na_position="last")
    pivot = pivot.reset_index(drop=True)
    if topn:
        pivot = pivot.head(topn)
    return pivot





@st.cache_data(show_spinner="데이터 처리 중… (첫 로드는 행 수에 따라 수십 초 걸릴 수 있어요)")
def load_upload(raw: bytes) -> pd.DataFrame:
    """원본 bytes → (원본이면 가공) → 최종 DataFrame.
    가공 경로는 xlsx 직렬화/재파싱을 생략하고 메모리에서 바로 DataFrame 을 만든다(대용량 속도 핵심)."""
    rows = _parse_raw_rows(raw)
    if _is_already_processed(rows):
        return _finalize_df(read_excel_smart(io.BytesIO(raw)))  # 완성형은 그대로
    col_names, data_rows = _process_raw_rows(rows)
    df = pd.DataFrame(data_rows, columns=col_names, dtype=object)
    return _finalize_df(df)


def _finalize_df(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [clean_col_name(c) for c in df.columns]

    # Detect columns
    date_col = find_col(df, ["출고날짜", "출고일", "판매일자", "주문일자"], "날짜")
    qty_col = find_col(df, ["수량", "판매수량"], "수량")
    gross_col = find_col(df, ["매출가", "총매출액", "매출액"], "매출")
    net_col = find_col(df, ["최종판매가", "순매출액", "실매출액"], "최종")
    profit_col = find_col(df, ["수익원(실배송비)", "수익원 실배송비", "수익원", "공헌이익"], "수익원")
    cost_col = find_col(df, ["원가총액", "출고원가"], "원가")
    mall_col = find_col(df, ["쇼핑몰", "몰", "채널"], "쇼핑몰")
    brand_col = find_col(df, ["브랜드", "브랜드명"], "브랜드")
    # 대분류: 위에서 AA열을 '대분류'로 보존했다. G열 '대카테고리'(브랜드패션 등)는 절대 쓰지 않는다.
    if "대분류" in df.columns:
        category_col = "대분류"
    else:
        category_col = find_col(df, ["카테고리", "분류"], "분류")
    model_col = find_col(df, ["모델명", "상품명", "품목명", "상품코드"], "모델")
    order_col = find_col(df, ["주문번호", "주문ID", "주문코드"], "주문")
    note_col = find_col(df, ["비고", "상태", "구분"], "비고")

    # 공식/병행 구분 컬럼: 헤더명 우선, 없으면 값이 공식/병행 으로만 이루어진 컬럼을 자동 탐색
    official_col = find_col(df, ["공식/병행", "공식병행", "공식여부"])
    if official_col is None:
        for _c in df.columns:
            _vals = set(df[_c].dropna().astype(str).str.strip().unique())
            if _vals and _vals <= {"공식", "병행"}:
                official_col = _c
                break

    # Standardize important columns
    if date_col is None:
        raise ValueError("날짜 컬럼을 찾지 못했습니다. '출고날짜' 또는 날짜가 포함된 컬럼이 필요합니다.")
    df["날짜"] = pd.to_datetime(df[date_col], errors="coerce")
    df = df.dropna(subset=["날짜"]).copy()
    # 반품(수량<0): 비고의 '반품(YYYY-MM-DD)' 실제 반품일을 출고날짜/분석날짜로 사용 (출고일 혼동 방지)
    if qty_col and note_col and qty_col in df.columns and note_col in df.columns:
        _ret = pd.to_numeric(df[qty_col], errors="coerce").fillna(0) < 0
        if _ret.any():
            _rd = pd.to_datetime(
                df.loc[_ret, note_col].astype(str).str.extract(r"(\d{4}[-./]\d{1,2}[-./]\d{1,2})", expand=False),
                errors="coerce")
            df.loc[_ret, "날짜"] = _rd.fillna(df.loc[_ret, "날짜"])
            if date_col and date_col in df.columns:
                df.loc[_ret, date_col] = df.loc[_ret, "날짜"]
    df["연도"] = df["날짜"].dt.year.astype(int)
    df["월"] = df["날짜"].dt.month.astype(int)
    df["연월"] = df["날짜"].dt.to_period("M").astype(str)
    # (주간/월간 기간 파생은 아래 수량 표준화 후에 수행 — 반품=수량 음수 제외 위해)
    df["요일순"] = df["날짜"].dt.weekday.astype(int)
    df["요일라벨"] = df["요일순"].map({0: "월", 1: "화", 2: "수", 3: "목", 4: "금", 5: "토", 6: "일"})

    for col in [qty_col, gross_col, net_col, profit_col, cost_col]:
        if col and col in df.columns:
            df[col] = to_number(df[col])

    if qty_col and qty_col != "수량":
        df["수량"] = df[qty_col]
    elif "수량" not in df.columns:
        df["수량"] = 0

    # (주간/월간/연간 기간 파생[주시작·주차]은 본문에서 '합쳐진 전체 df' 기준으로 수행한다.
    #  멀티파일 업로드 시 파일마다 따로 잡혀 라벨/모드가 틀어지는 문제 방지)

    if gross_col and gross_col != "매출가":
        df["매출가"] = df[gross_col]
    elif "매출가" not in df.columns:
        df["매출가"] = 0

    if net_col and net_col != "최종판매가":
        df["최종판매가"] = df[net_col]
    elif "최종판매가" not in df.columns:
        df["최종판매가"] = df["매출가"]

    # 매출가 컬럼이 비어있는(거의 0인) export 에서는 최종판매가를 매출가로 사용한다.
    if df["매출가"].abs().sum() == 0 or (df["매출가"] != 0).mean() < 0.05:
        df["매출가"] = df["최종판매가"]

    if profit_col and profit_col != "수익원(실배송비)":
        df["수익원(실배송비)"] = df[profit_col]
    elif "수익원(실배송비)" not in df.columns:
        df["수익원(실배송비)"] = 0

    if cost_col and cost_col != "원가총액":
        df["원가총액"] = df[cost_col]
    elif "원가총액" not in df.columns:
        df["원가총액"] = 0

    for std, col in {
        "쇼핑몰": mall_col,
        "브랜드": brand_col,
        "대분류": category_col,
        "공식/병행": official_col,
        "모델명": model_col,
        "주문번호": order_col,
        "비고": note_col,
    }.items():
        if col and col in df.columns:
            df[std] = df[col].fillna("미분류").astype(str)
        elif std not in df.columns:
            df[std] = "미분류"
        else:
            df[std] = df[std].fillna("미분류").astype(str)

    # Normalize text values
    for c in ["쇼핑몰", "브랜드", "대분류", "공식/병행", "모델명", "비고"]:
        df[c] = df[c].replace({"nan": "미분류", "None": "미분류", "": "미분류"})

    return df


@st.cache_data(show_spinner=False)
def _img_keys(name) -> set:
    """이미지 매칭 키 후보: 원본 + 괄호 사이즈 제거형(카드의 to_line 매칭용)."""
    s = "" if name is None else str(name).strip()
    out = {s}
    out.add(re.sub(r"\s*\([^()]*\)\s*$", "", s).strip())
    return {k for k in out if k}


@st.cache_data(show_spinner=False)
def load_image_map_from_bytes(data: bytes | None) -> dict:
    """엑셀의 두 번째(옆) 시트에서 A열=라인명, B열=이미지URL 매핑을 읽는다.
    시트가 없거나 형식이 안 맞으면 빈 dict 를 반환(이미지 없이 동작)."""
    if data is None:
        return {}
    try:
        xls = pd.ExcelFile(io.BytesIO(data))
    except Exception:
        return {}
    other_sheets = list(xls.sheet_names[1:])  # 첫 시트(메인 데이터) 제외
    mapping: dict[str, str] = {}
    for sh in other_sheets:
        try:
            sub = pd.read_excel(xls, sheet_name=sh, header=None, usecols=[0, 1])
        except Exception:
            continue
        for _, row in sub.iterrows():
            name = row.iloc[0]
            url = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            if pd.isna(name) or not url.lower().startswith(("http://", "https://")):
                continue  # 헤더행/빈값/비 URL 스킵
            for k in _img_keys(name):
                mapping[k] = url
    return mapping


def _norm_mall(s) -> str:
    """쇼핑몰명 매칭용 정규화: 공백/괄호/'주식회사' 제거, 소문자."""
    s = "" if s is None else str(s)
    s = re.sub(r"\s+", "", s).lower()
    for t in ("주식회사", "(", ")", "（", "）"):
        s = s.replace(t, "")
    return s


def load_targets_from_file(path) -> pd.DataFrame:
    """이미지 엑셀의 '이미지'가 아닌 시트(목표매출)에서 (공식/병행, 쇼핑몰, 월)별 목표 파싱.
    레이아웃: 'N월' 헤더 + 그 왼쪽 칸이 쇼핑몰명, 블록 라벨('병행'/'공식'), 합계행(쇼핑몰칸='쇼핑몰')."""
    cols = ["공식병행", "쇼핑몰", "월", "목표", "_key"]
    if path is None:
        return pd.DataFrame(columns=cols)
    try:
        xls = pd.ExcelFile(path)
    except Exception:
        return pd.DataFrame(columns=cols)
    sheets = [s for s in xls.sheet_names if str(s).strip() != "이미지"]
    if not sheets:
        return pd.DataFrame(columns=cols)
    try:
        raw = pd.read_excel(xls, sheet_name=sheets[0], header=None)
    except Exception:
        return pd.DataFrame(columns=cols)
    rows = raw.values.tolist()
    recs = []
    cur_block = None
    month_cols: dict[int, int] = {}
    for r in rows:
        cells = list(r)
        # 블록 헤더: 앞쪽 칸에 '병행'/'공식' + 행에 'N월' 라벨 존재
        blk = None
        for c in cells[:3]:
            cs = str(c).strip() if c is not None else ""
            if cs in ("병행", "공식"):
                blk = cs
                break
        has_month = any(c is not None and re.fullmatch(r"\d{1,2}월", str(c).strip()) for c in cells)
        if blk and has_month:
            cur_block = blk
            month_cols = {}
            for ci, c in enumerate(cells):
                m = re.fullmatch(r"(\d{1,2})월", str(c).strip()) if c is not None else None
                if m:
                    month_cols[ci] = int(m.group(1))
            continue
        if cur_block is None or not month_cols:
            continue
        name_idx = min(month_cols) - 1  # 쇼핑몰명은 월 컬럼 바로 왼쪽
        name = cells[name_idx] if 0 <= name_idx < len(cells) else None
        nm = str(name).strip() if name is not None else ""
        if (not nm) or nm.lower() == "nan" or nm in ("쇼핑몰", "목표매출"):
            continue  # 합계행/헤더/빈행
        for ci, mon in month_cols.items():
            v = pd.to_numeric(cells[ci], errors="coerce") if ci < len(cells) else np.nan
            if pd.notna(v) and v != 0:
                recs.append({"공식병행": cur_block, "쇼핑몰": nm, "월": int(mon),
                             "목표": float(v), "_key": _norm_mall(nm)})
    return pd.DataFrame(recs, columns=cols)


def find_image_file() -> Path | None:
    """앱 폴더에서 독립 이미지 매핑 파일('이미지*.xlsx' 등)을 찾는다(최신 우선)."""
    for pat in ("이미지*.xlsx", "이미지*.xls", "image*.xlsx", "images*.xlsx"):
        c = sorted(APP_DIR.glob(pat), key=lambda p: (p.stat().st_mtime, p.name), reverse=True)
        if c:
            return c[0]
    return None


def find_raw_files() -> list[Path]:
    """앱 폴더에서 매출 raw 파일을 자동으로 찾는다.
    'n월n주'(예: 10월3주차.xlsx) 패턴을 우선 사용하고, 없으면 이미지 파일을 제외한
    나머지 엑셀을 raw 로 본다. 여러 개면 모두 반환(병합) — 최신 수정 우선."""
    cand: dict[str, Path] = {}
    for pat in ("*.xlsx", "*.xls"):
        for p in APP_DIR.glob(pat):
            nm = p.name
            if nm.startswith("~$") or nm.startswith("."):       # 엑셀 임시/숨김 파일 제외
                continue
            low = nm.lower()
            if low.startswith(("이미지", "image", "images")):    # 이미지 매핑 파일 제외
                continue
            cand[str(p)] = p
    files = list(cand.values())
    # 'n월 n주' 패턴 파일을 우선. 하나라도 있으면 그 파일들만 raw 로 사용.
    wk = [p for p in files if re.search(r"\d+\s*월", p.name) and "주" in p.name]
    use = wk if wk else files
    use.sort(key=lambda p: (p.stat().st_mtime, p.name), reverse=True)
    return use


def load_image_map_from_file(path) -> dict:
    """독립 '이미지' 엑셀(첫 시트): A열=라인명/모델명, B열=이미지URL → {키: url}."""
    if path is None:
        return {}
    try:
        df = pd.read_excel(path, sheet_name=0, header=None, usecols=[0, 1])
    except Exception:
        return {}
    mapping: dict[str, str] = {}
    for _, row in df.iterrows():
        name = row.iloc[0]
        url = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        if pd.isna(name) or not url.lower().startswith(("http://", "https://")):
            continue
        for k in _img_keys(name):
            mapping[k] = url
    return mapping


@st.cache_data(show_spinner=False)
def load_image_map_from_image_xlsx(data: bytes | None) -> dict:
    """독립 '이미지' 엑셀에서 이미지 매핑을 읽는다. 시트명에 '이미지'가 있으면 그 시트,
    없으면 첫 시트의 A열=라인명/모델명, B열=이미지URL → {키: url}."""
    if data is None:
        return {}
    try:
        xls = pd.ExcelFile(io.BytesIO(data))
    except Exception:
        return {}
    target = None
    for s in xls.sheet_names:
        if "이미지" in str(s) or "image" in str(s).lower():
            target = s
            break
    if target is None:
        target = xls.sheet_names[0]
    try:
        df = pd.read_excel(xls, sheet_name=target, header=None, usecols=[0, 1])
    except Exception:
        return {}
    mapping: dict[str, str] = {}
    for _, row in df.iterrows():
        name = row.iloc[0]
        url = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        if pd.isna(name) or not url.lower().startswith(("http://", "https://")):
            continue
        for k in _img_keys(name):
            mapping[k] = url
    return mapping





@st.cache_data(show_spinner=False)
def load_line_map(src) -> dict:
    """엑셀에서 모델명→라인명 매핑을 읽는다(라인명 '전용' 시트가 있을 때만).
    시트명에 '라인'/'line'/'매핑'이 들어간 시트만 사용하고, 헤더에서 '모델명'·'라인명' 열을
    찾아 매핑한다(열 순서 무관). 헤더가 없으면 A열=모델명, B열=라인명으로 본다.
    ※ '재고'/'목표' 같은 시트는 여기서 읽지 않는다(재고는 load_stock 이 따로 처리)."""
    if src is None:
        return {}
    try:
        xls = pd.ExcelFile(io.BytesIO(src)) if isinstance(src, (bytes, bytearray)) else pd.ExcelFile(src)
    except Exception:
        return {}
    target = None
    for s in xls.sheet_names:
        low = str(s).lower().strip()
        if any(k in low for k in ("라인", "line", "매핑")):
            target = s
            break
    if target is None:
        return {}
    try:
        raw = pd.read_excel(xls, sheet_name=target, header=None)
    except Exception:
        return {}
    rows = [list(r) for r in raw.where(pd.notna(raw), None).values.tolist()]
    if not rows:
        return {}
    mcol = lcol = None
    hdr_row = None
    for i in range(min(5, len(rows))):
        for j, c in enumerate(rows[i]):
            cs = str(c).replace(" ", "").replace("\n", "") if c is not None else ""
            if cs in ("모델명", "모델", "상품코드", "상품명") and mcol is None:
                mcol, hdr_row = j, i
            if cs in ("라인명", "라인") and lcol is None:
                lcol = j
        if mcol is not None and lcol is not None:
            break
    m: dict[str, str] = {}
    if mcol is not None and lcol is not None:
        for r in rows[(hdr_row or 0) + 1:]:
            if mcol < len(r) and lcol < len(r):
                a, b = r[mcol], r[lcol]
                ka = _norm_model(a)
                vb = str(b).strip() if b is not None else ""
                if ka and vb and ka.lower() not in ("nan", "none") and vb.lower() not in ("nan", "none"):
                    m[ka] = vb
    else:  # 헤더 못 찾음 → A열=모델명, B열=라인명 가정
        for r in rows:
            if len(r) >= 2:
                a, b = r[0], r[1]
                ka = _norm_model(a)
                vb = str(b).strip() if b is not None else ""
                if not ka or ka in ("모델명", "모델", "상품명") or vb in ("라인명", "라인", ""):
                    continue
                if ka.lower() in ("nan", "none") or vb.lower() in ("nan", "none"):
                    continue
                m[ka] = vb
    return m


_STOCK_HDR_KEYS = ("라인명", "브랜드", "모델명", "수량", "총원가", "원가평균", "가용수량")


def _xlsx_rows_fast_sheet(data: bytes, sheet_name) -> list:
    """openpyxl read_only 로 특정 시트만 값으로 빠르게 읽는다(대용량 재고 시트용)."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    while rows and all(c is None for c in rows[-1]):
        rows.pop()
    return rows


def _stock_rows_to_df(rows: list) -> pd.DataFrame:
    """재고 행 리스트 → 표준 재고 DataFrame[라인명,브랜드,모델명,수량,가용수량,원가평균,총원가].
    제목/숫자 행이 위에 있어도 헤더(라인명·브랜드·총원가…)를 자동 탐지한다."""
    if not rows:
        return pd.DataFrame()
    hdr_i = 0
    for i in range(min(8, len(rows))):
        vals = set(str(x).replace("\n", " ").strip() for x in rows[i] if x is not None)
        hit = sum(1 for k in _STOCK_HDR_KEYS if any(k == v or k in v for v in vals))
        if hit >= 4:
            hdr_i = i
            break
    header = _dedupe_cols([str(c).replace("\n", " ").strip() if c is not None else "" for c in rows[hdr_i]])
    sdf = pd.DataFrame(rows[hdr_i + 1:], columns=header)

    def col(*names):
        for name in names:  # 정확 매칭 우선
            key = name.replace(" ", "").replace("\n", "")
            for c in sdf.columns:
                if str(c).replace(" ", "").replace("\n", "") == key:
                    return c
        for name in names:  # 부분 매칭
            key = name.replace(" ", "").replace("\n", "")
            for c in sdf.columns:
                if key in str(c).replace(" ", "").replace("\n", ""):
                    return c
        return None

    c_line, c_brand, c_model = col("라인명"), col("브랜드"), col("모델명")
    c_qty, c_avail = col("수량"), col("가용수량")
    c_cost, c_total = col("원가평균"), col("총원가")
    c_daecat, c_cat = col("대카테고리"), col("카테고리")
    out = pd.DataFrame()
    out["라인명"] = sdf[c_line].astype(str).str.strip() if c_line else ""
    out["브랜드"] = sdf[c_brand].astype(str).str.strip() if c_brand else ""
    out["모델명"] = sdf[c_model].astype(str).str.strip() if c_model else out["라인명"]
    out["대카테고리"] = sdf[c_daecat].astype(str).str.strip() if c_daecat else ""
    out["카테고리"] = sdf[c_cat].astype(str).str.strip() if c_cat else ""
    out["수량"] = to_number(sdf[c_qty]) if c_qty else 0
    out["가용수량"] = to_number(sdf[c_avail]) if c_avail else np.nan
    out["원가평균"] = to_number(sdf[c_cost]) if c_cost else np.nan
    out["총원가"] = to_number(sdf[c_total]) if c_total else (out["수량"] * out["원가평균"])
    out["공식/병행"] = [
        _classify_official(b, d, c) for b, d, c in zip(out["브랜드"], out["대카테고리"], out["카테고리"])
    ]
    out = out[~(out["라인명"].isin(["", "nan", "None"]) & out["모델명"].isin(["", "nan", "None"]))]
    return out.reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_stock(src) -> pd.DataFrame:
    """행사가관리/재고 엑셀(첫 시트) → 표준 재고 DataFrame."""
    if src is None:
        return pd.DataFrame()
    try:
        data = bytes(src) if isinstance(src, (bytes, bytearray)) else Path(src).read_bytes()
    except Exception:
        return pd.DataFrame()
    try:
        rows = _xlsx_rows_fast(data)
    except Exception:
        try:
            raw = pd.read_excel(io.BytesIO(data), header=None)
            rows = [list(r) for r in raw.where(pd.notna(raw), None).values.tolist()]
        except Exception:
            return pd.DataFrame()
    return _stock_rows_to_df(rows)


@st.cache_data(show_spinner=False)
def load_stock_from_image_xlsx(src) -> pd.DataFrame:
    """이미지 엑셀에 '재고'/'stock' 시트가 있으면 그 시트로 재고를 읽는다
    (별도 재고 파일을 올리지 않아도 됨)."""
    if src is None:
        return pd.DataFrame()
    try:
        data = bytes(src) if isinstance(src, (bytes, bytearray)) else Path(src).read_bytes()
    except Exception:
        return pd.DataFrame()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True)
        names = list(wb.sheetnames)
        wb.close()
    except Exception:
        return pd.DataFrame()
    target = None
    for s in names:
        if any(k in str(s).lower() for k in ("재고", "stock", "event_price", "행사가")):
            target = s
            break
    if target is None:
        return pd.DataFrame()
    try:
        rows = _xlsx_rows_fast_sheet(data, target)
    except Exception:
        return pd.DataFrame()
    return _stock_rows_to_df(rows)


def line_map_from_stock(stock_df: pd.DataFrame) -> dict:
    """재고 DataFrame 에서 모델명→라인명 매핑 추출(이미지 엑셀 매핑 보조)."""
    if stock_df is None or stock_df.empty:
        return {}
    m: dict[str, str] = {}
    for md, ln in zip(stock_df["모델명"], stock_df["라인명"]):
        k, v = _norm_model(md), str(ln).strip()
        if k and v and k.lower() not in ("nan", "none", "") and v.lower() not in ("nan", "none", ""):
            m.setdefault(k, v)
    return m


# ============================================================
# 원본(raw) 업로드 자동 전처리  ── process_excel.py 로직 포팅(값 전용)
#   HTML 위장 xls / 진짜 xlsx·xls 자동 인식 → 사은품 배송비 이전,
#   행 삭제, 쇼핑몰명 통일, 공식/병행 분류, 정산금·대분류 추가.
#   이미 가공된(완성) 파일이면 그대로 통과.
# ============================================================
# 열 인덱스(0-based) — process_excel.py 와 동일
_CI, _DI, _EI, _FI, _GI, _HI, _WI = 2, 3, 4, 5, 6, 8, 23
_CATEGORY_IDX, _M_IDX, _O_IDX = 7, 12, 14
_P_IDX, _Q_IDX, _ORDER_IDX, _BRAND_IDX = 15, 16, 1, 5
_TRUNCATE_IDX = 25
_NUM_COL_RANGE = range(9, 23)
_GIFT_BRANDS = ["쇼핑백", "사은품"]
_DELETE_H_VALUES = {"파슬AS", "쿠팡그로스 재고손실보상", "쿠팡그로스 기타정산"}
_DELETE_D_KEYWORDS = ["방송", "홈방", "나린인터", "태그바이"]
_D_RENAME_MAP = {
    "KREAM": "크림 주식회사", "카카오톡선물하기_디젤": "카카오톡선물하기",
    "카카오톡선물하기_병행": "카카오톡선물하기", "카카오톡선물하기_공식": "카카오톡선물하기",
    "에이블리(블리블리)": "에이블리", "에이블리(치페)": "에이블리",
    "무신사_블리블리": "무신사", "Wconcept(뷰티)": "Wconcept",
    "롯데백화점 온라인몰 공식": "롯데백화점온라인몰",
    "29CM(티켓투더문)": "29CM(공식)", "29CM(디젤)": "29CM(공식)",
    "카카오스타일 (치페)": "카카오스타일 (지그재그)",
    "카카오스타일 (티켓투더문)": "카카오스타일 (지그재그)",
    "카카오스타일 (블리블리)": "카카오스타일 (지그재그)",
}
_OFFICIAL_F_ONLY = {
    "블리블리", "헤브블루", "미스그린", "치페", "파슬", "아르마니", "티켓투더문",
    "아르마니익스체인지", "울프1834", "인도솔", "썬젤리", "스카겐", "미니쿄모", "스케쳐스",
}
_CATEGORY_MAP = {
    "가방": "가방", "귀걸이": "주얼리", "드레스": "의류", "라이터": "용품", "마사지볼": "용품",
    "모자": "소품", "목걸이": "주얼리", "문구": "용품", "반지": "주얼리", "밴드": "시계",
    "벨트": "소품", "상의": "의류", "시계": "시계", "신발": "신발", "아우터": "의류",
    "잡화ACC": "소품", "지갑": "지갑", "침낭": "용품", "키링&키홀더": "소품", "팔찌": "주얼리",
    "폼롤러": "용품", "하의": "의류", "핸드폰케이스": "소품", "홈데코": "용품", "우산": "소품",
    "옷걸이": "용품", "에어팟케이스": "용품", "언더웨어": "의류", "바디케어": "용품",
    "쇼핑백": "용품", "향수": "용품", "스킨케어": "용품", "거치대": "시계", "인솔": "용품",
    "쥬얼리보관함": "주얼리", "와인더": "시계", "시계보관함": "시계", "완구": "용품",
    "손난로": "용품", "참": "주얼리", "보온주머니": "용품", "생활잡화": "용품",
    "스포츠용품": "용품", "스윔웨어": "용품", "수납용품": "용품", "브로치": "소품",
    "케이블": "시계", "생활용품": "용품", "욕실용품": "용품", "슬립웨어": "의류",
    "아이메이크업": "용품", "립메이크업": "용품", "베이스메이크업": "용품", "뷰티소품": "용품",
    "클렌징": "용품", "선케어": "용품", "헤어케어": "용품", "주방용품": "용품",
}
_HDR_KEYWORDS = {"주문번호", "쇼핑몰", "브랜드", "수량", "최종판매가", "출고날짜"}
# 완성 파일의 컬럼명(위치 0~24). process_excel 이 위치(인덱스)로 처리하므로
# 원본 헤더명이 무엇이든 출력은 이 표준명을 위치 기준으로 박아 대시보드가 항상 인식하게 한다.
_CANON_HEADERS = [
    "차수", "주문번호", "품목코드", "쇼핑몰", "쇼핑몰아이디", "브랜드", "대카테고리", "카테고리",
    "모델명", "수량", "판매단가", "매출가", "최종판매가", "수수료", "수수료액", "마켓설정배송비",
    "실배송비", "출고원가", "원가총액", "수익원(마켓설정배송비)", "수익율(마켓설정배송비)",
    "수익원(실배송비)", "수익율(실배송비)", "출고날짜", "비고",
]


def _rawcell(row, idx):
    try:
        v = row[idx]
        return "" if v is None else str(v).strip()
    except (IndexError, KeyError):
        return ""


def _to_num(s):
    try:
        return float(str(s).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _get_d(row):
    d = _rawcell(row, _DI); e = _rawcell(row, _EI)
    try:
        if int(float(e)) == 1039456 and d.upper() == "GS SHOP":
            return "GS_API"
    except (ValueError, TypeError):
        pass
    if e == "033139LT":
        return "롯데홈쇼핑_API"
    return _D_RENAME_MAP.get(d, d)


def _classify_official(brand, daecat, cat) -> str:
    """브랜드 + 대카테고리 + 카테고리로 공식/병행 분류 (판매·재고 공통 로직)."""
    f = str(brand).strip(); g = str(daecat).strip(); h = str(cat).strip()
    if f in _OFFICIAL_F_ONLY:                               return "공식"
    if f == "마이클코어스" and g == "시계쥬얼리":          return "공식"
    if f == "디젤" and g == "시계쥬얼리":                  return "공식"
    if f == "라코스테" and g == "브랜드패션":              return "공식"
    if f == "토리버치" and (h.startswith("TBW") or h == "시계"):  return "공식"
    if f == "비비안웨스트우드" and h.startswith("VV"):     return "공식"
    return "병행"


def _get_c(row):
    return _classify_official(_rawcell(row, _FI), _rawcell(row, _GI), _rawcell(row, _HI))


def _should_delete(row):
    if _rawcell(row, _HI) in _DELETE_H_VALUES:
        return True
    return any(kw in _rawcell(row, _DI) for kw in _DELETE_D_KEYWORDS)


def _parse_raw_rows(data: bytes) -> list:
    """원본 파일 → 행 리스트(값). HTML 위장 / 진짜 xlsx 자동 인식. 무거운 파일도 빠르게."""
    head = data[:512].lower().lstrip()
    is_html = head[:1] == b"<" or b"<table" in head or b"<html" in head or b"<meta" in head
    if is_html:
        text = None
        for enc in ("utf-8", "cp949", "euc-kr"):
            try:
                text = data.decode(enc); break
            except Exception:
                text = None
        if text is None:
            text = data.decode("utf-8", errors="ignore")
        try:
            rows = _html_rows_fast(text)          # lxml 직접(빠름)
        except Exception:
            raw = pd.read_html(io.StringIO(text), header=None)[0]
            rows = [list(r) for r in raw.where(pd.notna(raw), None).values.tolist()]
    else:
        try:
            rows = _xlsx_rows_fast(data)          # openpyxl read_only(값만, 무거운 파일도 빠름)
        except Exception:
            raw = pd.read_excel(io.BytesIO(data), header=None, dtype=object)  # 구형 .xls(xlrd) 폴백
            rows = [list(r) for r in raw.where(pd.notna(raw), None).values.tolist()]
    return rows


def _is_already_processed(rows) -> bool:
    flat = set()
    for r in rows[:5]:
        for c in r:
            flat.add(str(c).strip() if c is not None else "")
    return "대분류" in flat or "정산금" in flat


def _find_header_idx(rows) -> int:
    """헤더(컬럼명) 행 인덱스. 없으면 -1(데이터가 0행부터 시작)."""
    for i in range(min(6, len(rows))):
        vals = set(str(x).strip() for x in rows[i] if x is not None)
        if len(vals & _HDR_KEYWORDS) >= 3:
            return i
    return -1


def _process_raw_rows(rows):
    """행 리스트 → (컬럼명, 데이터행) : 사은품 이전·삭제·분류·정산금/대분류."""
    hdr = _find_header_idx(rows)
    data_start = hdr + 1 if hdr >= 0 else 0  # 헤더 없으면 0행부터
    data_rows = [list(r) for r in rows[data_start:] if any(str(c).strip() for c in r if c is not None)]

    def is_gift(r):
        return _rawcell(r, _BRAND_IDX) in _GIFT_BRANDS

    # 사은품/쇼핑백 배송비를 같은 주문의 정상상품 첫 행으로 이전
    order_to_normal = {}
    for i, r in enumerate(data_rows):
        if not is_gift(r):
            order_to_normal.setdefault(_rawcell(r, _ORDER_IDX), []).append(i)
    orphan = set()
    for i, r in enumerate(data_rows):
        if not is_gift(r):
            continue
        targets = order_to_normal.get(_rawcell(r, _ORDER_IDX), [])
        if not targets:
            orphan.add(i); continue
        t = data_rows[targets[0]]
        for IDX in (_P_IDX, _Q_IDX):
            if IDX < len(r) and IDX < len(t):
                t[IDX] = _to_num(t[IDX]) + _to_num(r[IDX])
    kept = []
    for i, r in enumerate(data_rows):
        if is_gift(r) and i not in orphan:
            continue  # 매칭된 사은품 삭제
        if i in orphan and _BRAND_IDX < len(r):
            r[_BRAND_IDX] = _rawcell(r, _BRAND_IDX) + " ⚠미매칭"
        kept.append(r)
    data_rows = kept

    # 행 삭제 조건
    data_rows = [r for r in data_rows if not _should_delete(r)]

    # 쇼핑몰명 통일(D) / 공식·병행 분류(C)
    for r in data_rows:
        if _DI < len(r):
            r[_DI] = _get_d(r)
        if _CI < len(r):
            r[_CI] = _get_c(r)

    # 컬럼명: 원본 헤더명 대신 '표준명(위치 기준)' 사용 → 대시보드가 항상 인식
    col_names = list(_CANON_HEADERS[:_TRUNCATE_IDX])
    while len(col_names) < _TRUNCATE_IDX:
        col_names.append(f"col{len(col_names)}")
    col_names = col_names + ["정산금", "대분류"]

    # 데이터: 정산금(=최종판매가-수수료액)·대분류 추가
    out = []
    for r in data_rows:
        m = _to_num(r[_M_IDX]) if _M_IDX < len(r) else 0.0
        o = _to_num(r[_O_IDX]) if _O_IDX < len(r) else 0.0
        cat = _rawcell(r, _CATEGORY_IDX)
        대분류 = _CATEGORY_MAP.get(cat, "")
        r = list(r[:_TRUNCATE_IDX])
        while len(r) < _TRUNCATE_IDX:
            r.append("")
        r.append(round(m - o, 2))                       # 정산금
        r.append(대분류 if (대분류 or not cat) else "❓미매핑")  # 대분류
        out.append(r)
    return col_names, out


def _rows_to_xlsx_bytes(col_names, data_rows) -> bytes:
    """가공 결과를 완성 구조 xlsx 바이트로 (숫자/날짜 타입 지정)."""
    from openpyxl import Workbook
    import datetime as _dt
    wb = Workbook(); ws = wb.active
    ws.append(col_names)
    settle_idx = len(col_names) - 2  # 정산금 위치
    date_fmts = ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%m/%d/%Y", "%d/%m/%Y"]
    for r in data_rows:
        row_out = []
        for ci, v in enumerate(r):
            if ci in _NUM_COL_RANGE or ci == settle_idx:
                row_out.append(_to_num(v))
            elif ci == _WI:
                if isinstance(v, (_dt.datetime, _dt.date, pd.Timestamp)):
                    row_out.append(pd.Timestamp(v).to_pydatetime())
                else:
                    s = str(v).strip()
                    s2 = s[:10] if (" " in s and len(s) >= 10) else s
                    dv = None
                    for fmt in date_fmts:
                        try:
                            dv = _dt.datetime.strptime(s2, fmt); break
                        except ValueError:
                            continue
                    row_out.append(dv if dv else s)
            else:
                row_out.append("" if v is None else v)
        ws.append(row_out)
    bio = io.BytesIO(); wb.save(bio)
    return bio.getvalue()


@st.cache_data(show_spinner=False)
def preprocess_upload(data: bytes) -> bytes:
    """원본이면 가공해 완성 구조 xlsx 바이트 반환. 이미 완성이면 원본 그대로."""
    try:
        rows = _parse_raw_rows(data)
    except Exception:
        return data  # 파싱 자체 실패 → 원본 그대로(로더가 직접 시도/에러)
    if not rows or _is_already_processed(rows):
        return data  # 이미 가공된(완성) 파일 → 그대로 통과
    col_names, out_rows = _process_raw_rows(rows)  # 실패 시 에러를 그대로 노출(원인 파악)
    return _rows_to_xlsx_bytes(col_names, out_rows)


# -----------------------------
# UI
# -----------------------------
st.title("📊 메카 매출 대시보드 (자동: 주간/월간/연간)")
st.caption("데이터 기간에 따라 주간(WoW)/월간(MoM)/연간(YoY) 자동 전환 · 공식/병행 총매출·수익률 · TOP 30")
st.info("🔧 build 트리플 v1 — 기간 ≤33일 주간(3주)·≤365일 월간(3개월)·>365일 연간(3개년) 자동 전환. 반품일 보정·목표 달성률·원본 자동가공·이미지 지원.", icon="ℹ️")

with st.sidebar:
    st.header("데이터")
    PRINT_MODE = st.checkbox("📄 인쇄/PDF용 보기", value=False,
                             help="켜면 표가 정적 표로 바뀌어 인쇄·PDF가 깔끔하게 나옵니다. (정렬 기능은 꺼짐)")
    if PRINT_MODE:
        st.caption("인쇄 모드: 브라우저 인쇄(Ctrl+P) 후 끄세요.")
    uploaded = st.file_uploader(
        "Excel 파일 업로드 (여러 개 선택 가능)",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
    )
    _imgf = find_image_file()
    if _imgf is not None:
        st.caption(f"🖼 이미지 매핑: **{_imgf.name}** (A열 라인명/모델명 · B열 이미지URL)")
    else:
        st.caption("🖼 상품 이미지: 같은 폴더에 '이미지.xlsx'(A열 라인명, B열 URL)를 두면 자동 표시됩니다.")
    st.divider()
    st.caption("📦 재고 현황: 이미지.xlsx 의 '재고' 시트를 읽어 맨 아래에 표시합니다.")
    _rawf = find_raw_files()
    if _rawf:
        _rawnames = ", ".join(p.name for p in _rawf[:5]) + (" 외" if len(_rawf) > 5 else "")
        st.caption(f"📂 폴더 raw 자동 로드: **{_rawnames}** (업로드하면 그게 우선)")
    else:
        st.caption("📂 자동 로드: 같은 폴더에 'n월n주' raw 파일(예: 10월3주차.xlsx)을 두면 업로드 없이 자동으로 열립니다.")

try:
    if uploaded:  # 수동 업로드가 최우선 (다중 허용 시 list)
        raw_sources = [(uf.name, uf.getvalue()) for uf in uploaded]
        _auto_loaded = False
    else:          # 업로드가 없으면 앱 폴더의 raw 파일을 자동으로 읽어 연다
        raw_sources = [(p.name, p.read_bytes()) for p in find_raw_files()]
        _auto_loaded = True
    if not raw_sources:
        st.info("왼쪽에서 매출 엑셀을 업로드하거나, 앱 폴더에 'n월n주' raw 파일(예: 10월3주차.xlsx)을 두면 자동으로 열립니다.", icon="📂")
        st.stop()
    raws = [b for _, b in raw_sources]
    frames = [load_upload(b) for b in raws]
    df = pd.concat(frames, ignore_index=True) if len(frames) > 1 else frames[0]
    img_map = {}
    for b in raws:  # 업로드/폴더 파일 안에 이미지 시트가 있으면 사용
        img_map.update(load_image_map_from_bytes(b))
    _names = ", ".join(n for n, _ in raw_sources)
    if _auto_loaded:
        st.success(f"📂 폴더에서 {len(raw_sources)}개 raw 파일 자동 로드 — {_names} · 총 {len(df):,}행", icon="✅")
    elif len(raw_sources) > 1:
        st.success(f"📎 {len(raw_sources)}개 파일 병합 분석 — {_names} · 총 {len(df):,}행", icon="✅")
    # 독립 '이미지' 엑셀 파일(앱 폴더의 이미지*.xlsx)이 있으면 병합 (우선 적용)
    _img_file = find_image_file()
    _img_bytes = _img_file.read_bytes() if _img_file is not None else None
    if _img_bytes is not None:
        img_map.update(load_image_map_from_image_xlsx(_img_bytes))
    targets_df = load_targets_from_file(_img_file)  # 목표매출(이미지 엑셀의 2번째 시트)
    # 모델명→라인명 매핑(라인명 전용 시트가 있을 때만) + 재고 로딩
    line_map = {}
    if _img_bytes is not None:
        line_map.update(load_line_map(_img_bytes))
    # 재고: 이미지.xlsx 의 '재고' 시트에서만 읽음
    stock_df = load_stock_from_image_xlsx(_img_bytes) if _img_bytes is not None else pd.DataFrame()
    if not stock_df.empty:
        line_map.update(line_map_from_stock(stock_df))  # 재고의 모델명→라인명
    # 판매 데이터에 라인명 부여(베스트 상품을 라인명으로 취합)
    df["라인명"] = df["모델명"].apply(_line_of) if "모델명" in df.columns else ""
except Exception as e:
    st.error(f"파일을 읽는 중 오류가 발생했습니다: {e}")
    st.stop()


def make_product_display(prod_df: pd.DataFrame, extra_cols: list, img_width: str = "small"):
    """집계된 상품 표(모델명 포함)에 라인명→이미지 매칭하여 (표시용 df, column_config) 반환.
    img_map 이 비어있으면 이미지 컬럼 없이 그대로 표시. img_width: small/medium/large."""
    out = prod_df.copy().reset_index(drop=True)
    if "라인명" not in out.columns:
        out["라인명"] = out["모델명"].apply(to_line) if "모델명" in out.columns else ""
    show_img = bool(img_map)
    if show_img:
        out["이미지"] = out["라인명"].map(img_map).fillna("")
    out.insert(0, "Rank", np.arange(1, len(out) + 1))
    cols = ["Rank"] + (["이미지"] if show_img else []) + extra_cols
    disp = format_table(out[cols])
    colcfg = {"Rank": st.column_config.TextColumn("#")}
    if show_img:
        colcfg["이미지"] = st.column_config.ImageColumn("이미지", width=img_width)
    return disp, colcfg


def product_cards_html(prod_df: pd.DataFrame, n: int = 10, img_px: int = 80, start: int = 1, step: int = 1) -> str:
    """상품을 카드형 HTML로 렌더 (이미지 크게, 순위는 인라인 #N).
    start/step: 순위 = start + i*step (좌우 교차 배치 시 step=2 등)."""
    rows = prod_df.head(n).reset_index(drop=True)
    show_imgs = bool(img_map)
    has_cat = "대분류" in rows.columns
    cards = []
    for i, r in rows.iterrows():
        rank = start + i * step
        brand = html.escape(str(r.get("브랜드", "")))
        cat = html.escape(str(r.get("대분류", ""))) if has_cat else ""
        meta = f"#{rank} · {brand}" + (f" · {cat}" if cat else "")
        line_name = str(r.get("라인명", "")).strip() or to_line(str(r.get("모델명", "")))
        model = html.escape(line_name if len(line_name) <= 38 else line_name[:37] + "…")
        rate = r.get("수익률", float("nan"))
        rate_s = f"{rate:.1f}%" if pd.notna(rate) and np.isfinite(rate) else "-"
        qty = r.get("수량", float("nan"))
        qty_s = f" · {int(qty):,}개" if pd.notna(qty) else ""
        sales_s = eok(r.get("최종판매가", 0))
        mall = str(r.get("최다몰", "")).strip()
        mall_pct = r.get("최다몰비중", float("nan"))
        if mall:
            _mpct = f" · {mall_pct:.0f}%" if pd.notna(mall_pct) and np.isfinite(mall_pct) else ""
            mall_block = (f'<div style="font-size:11px;color:#475569;margin-top:2px;">'
                          f'🏆 {html.escape(mall)}<span style="color:#94a3b8;">{_mpct}</span></div>')
        else:
            mall_block = ""
        img_block = ""
        if show_imgs:
            url = img_map.get(line_name, "")
            if url:
                img_block = (
                    f'<img src="{html.escape(url, quote=True)}" '
                    f'style="width:{img_px}px;height:{img_px}px;object-fit:cover;border-radius:8px;'
                    f'flex:0 0 auto;background:#f1f5f9;border:1px solid #eef2f7;">'
                )
            else:
                img_block = (
                    f'<div style="width:{img_px}px;height:{img_px}px;border-radius:8px;background:#f1f5f9;'
                    f'flex:0 0 auto;display:flex;align-items:center;justify-content:center;'
                    f'color:#cbd5e1;font-size:10px;">no img</div>'
                )
        cards.append(
            f'<div style="display:flex;gap:10px;align-items:center;padding:8px 8px;'
            f'border-bottom:1px solid #eef2f7;">{img_block}'
            f'<div style="min-width:0;flex:1;">'
            f'<div style="font-size:11px;color:#94a3b8;">{meta}</div>'
            f'<div style="font-size:13px;font-weight:600;color:#0f172a;white-space:nowrap;'
            f'overflow:hidden;text-overflow:ellipsis;">{model}</div>'
            f'<div style="font-size:13px;color:#0f172a;">{sales_s} '
            f'<span style="color:#64748b;">· {rate_s}{qty_s}</span></div>'
            f'{mall_block}'
            f'</div></div>'
        )
    return (
        '<div style="border:1px solid #e8eef5;border-radius:12px;overflow:hidden;'
        'box-shadow:0 2px 8px rgba(15,23,42,0.03);">' + "".join(cards) + "</div>"
    )


VIBRANT_COLORS = [
    "#2563eb", "#f59e0b", "#10b981", "#ec4899", "#8b5cf6",
    "#06b6d4", "#ef4444", "#eab308", "#6366f1", "#14b8a6",
]


def share_donut(agg_df: pd.DataFrame, name_col: str, value_col: str, title: str, cmap: dict | None = None):
    """비중 도넛(생기있는 색). 슬라이스 라벨 = 이름 + 비중%만(크게). 값이 양수인 항목만."""
    d = agg_df.copy()
    d = d[pd.to_numeric(d[value_col], errors="coerce").fillna(0) > 0]
    names = d[name_col].astype(str).tolist()
    vals = list(d[value_col])
    if cmap:
        colors = [cmap.get(n, "#9aa5b1") for n in names]
    else:
        colors = [VIBRANT_COLORS[i % len(VIBRANT_COLORS)] for i in range(len(names))]
    fig = go.Figure(
        go.Pie(
            labels=names, values=vals, hole=0.5,
            textinfo="label+percent", textposition="inside",
            insidetextorientation="horizontal", textfont=dict(size=15),
            marker=dict(colors=colors, line=dict(color="#ffffff", width=2)),
            sort=False, direction="clockwise",
        )
    )
    fig.update_layout(title=title, showlegend=False, margin=dict(t=50, b=10, l=10, r=10),
                      font=dict(size=14), yaxis=dict(tickformat=","))
    return fig

# Sidebar filters
with st.sidebar:
    st.header("필터")
    # 주차(시간순) — 최근 3구간만 사용하고 그 이전 데이터는 컷한다.
    RECENT_WEEKS = 3
    # 데이터 기간으로 주간/월간/연간 모드 결정 (load 와 동일: 반품=수량 음수 제외)
    #   33일 이하 → 주간(3주) / 365일 이하 → 월간(3개월) / 365일 초과 → 연간(3개년)
    _qn = pd.to_numeric(df["수량"], errors="coerce").fillna(0) if "수량" in df.columns else None
    _dts = df.loc[_qn >= 0, "날짜"].dropna() if _qn is not None else df["날짜"].dropna()
    if _dts.empty:
        _dts = df["날짜"].dropna()
    _span = int((_dts.max() - _dts.min()).days) if len(_dts) else 0
    if _span > 365:
        MODE = "yearly"
        THIS, PREV, WOW, INTERVAL_LABEL = "올해", "전년", "YoY", "연간"
        PERIODLY, PERIOD_CMP, TREND_N, PERIOD_AXIS = "연도별", "3개년 비교", "3개년", "연도"
        RECENT_TREND_TITLE = "최근 3개년 매출 추이"
        _filter_label = f"표시 연도 (최근 {RECENT_WEEKS}개년)"
        _mode_note = f"데이터 기간 {_span}일 → 📅 **연간(YoY)** 모드 · 최근 3개년"
    elif _span > 33:
        MODE = "monthly"
        THIS, PREV, WOW, INTERVAL_LABEL = "이번달", "전월", "MoM", "월간"
        PERIODLY, PERIOD_CMP, TREND_N, PERIOD_AXIS = "월별", "3개월 비교", "3개월", "월"
        RECENT_TREND_TITLE = "최근 3개월 매출 추이"
        _filter_label = f"표시 월 (최근 {RECENT_WEEKS}개월)"
        _mode_note = f"데이터 기간 {_span}일 → 📅 **월간(MoM)** 모드 · 최근 3개월"
    else:
        MODE = "weekly"
        THIS, PREV, WOW, INTERVAL_LABEL = "이번주", "전주", "WoW", "주간"
        PERIODLY, PERIOD_CMP, TREND_N, PERIOD_AXIS = "주차별", "3주 비교", "3주", "주차"
        RECENT_TREND_TITLE = "최근 3주 매출 추이"
        _filter_label = f"표시 주차 (최근 {RECENT_WEEKS}주)"
        _mode_note = f"데이터 기간 {_span}일 → 📅 **주간(WoW)** 모드 · 최근 3주"
    MONTHLY = MODE != "weekly"  # (기존 호환용)
    # 기간 파생(주시작/주차)을 '합쳐진 전체 df' 기준으로 — 멀티파일도 일관되게
    if MODE == "yearly":
        df["주시작"] = df["날짜"].dt.to_period("Y").dt.to_timestamp()  # 그 해 1월 1일
        df["주차"] = df["주시작"].dt.year.astype(int).astype(str) + "년"
    elif MODE == "monthly":
        df["주시작"] = df["날짜"].dt.to_period("M").dt.to_timestamp()  # 그 달 1일
        df["주차"] = df["주시작"].dt.year.astype(int).astype(str) + "년 " + df["주시작"].dt.month.astype(int).astype(str) + "월"
    else:
        df["주시작"] = (df["날짜"] - pd.to_timedelta(df["날짜"].dt.weekday, unit="D")).dt.normalize()
        _mon = df["주시작"]
        _fw = _mon.dt.to_period("M").dt.start_time.dt.weekday
        _wom = ((_mon.dt.day + _fw - 1) // 7 + 1).astype(int)
        df["주차"] = _mon.dt.month.astype(int).astype(str) + "월 " + _wom.astype(str) + "주차"
    _wk = df[["주시작", "주차"]].dropna().drop_duplicates().sort_values("주시작")
    week_order_all = _wk["주차"].tolist()[-RECENT_WEEKS:]
    df_all = df.copy()  # 컷 전 전체(이번달 총 달성률 계산용)
    df = df[df["주차"].isin(week_order_all)].copy()
    selected_weeks = st.multiselect(_filter_label, week_order_all, default=week_order_all)

    def multiselect_all(label, options, max_default=20):
        options = sorted([x for x in options if pd.notna(x)])
        return st.multiselect(label, options, default=options)

    selected_malls = multiselect_all("쇼핑몰", df["쇼핑몰"].unique())
    selected_brands = multiselect_all("브랜드", df["브랜드"].unique())
    selected_types = multiselect_all("공식/병행", df["공식/병행"].unique())
    selected_cats = multiselect_all("대분류", df["대분류"].unique())
    selected_notes = multiselect_all("비고", df["비고"].unique())

    include_returns = st.checkbox("반품/음수 데이터 포함", value=True)

f = df[
    df["주차"].isin(selected_weeks)
    & df["쇼핑몰"].isin(selected_malls)
    & df["브랜드"].isin(selected_brands)
    & df["공식/병행"].isin(selected_types)
    & df["대분류"].isin(selected_cats)
    & df["비고"].isin(selected_notes)
].copy()

# 화면에서 사용할 주차(시간순) 목록
week_order = [w for w in week_order_all if w in set(f["주차"].unique())]

if not include_returns:
    f = f[(f["수량"] >= 0) & (f["최종판매가"] >= 0)].copy()

if f.empty:
    st.warning("필터 조건에 해당하는 데이터가 없습니다.")
    st.stop()

metric_cols = {
    "최종판매가": "최종판매가",
    "수량": "수량",
    "수익원(실배송비)": "수익원(실배송비)",
}

# -----------------------------
# KPI
# -----------------------------
latest_week = week_order[-1] if week_order else None
prev_week = week_order[-2] if len(week_order) >= 2 else None
fw = f[f["주차"] == latest_week].copy() if latest_week else f.copy()

latest_sales = fw["최종판매가"].sum()
prev_sales = f[f["주차"] == prev_week]["최종판매가"].sum() if prev_week else 0
wow_rate = ((latest_sales - prev_sales) / abs(prev_sales) * 100) if prev_sales else np.nan

now_qty = fw["수량"].sum()
now_profit = fw["수익원(실배송비)"].sum()
now_avg = latest_sales / now_qty if now_qty else 0
now_rate = now_profit / latest_sales * 100 if latest_sales else 0
order_count = fw["주문번호"].nunique() if "주문번호" in fw.columns else len(fw)

st.caption(_mode_note)
cols = st.columns(5)
cols[0].metric(f"{THIS} 매출 ({latest_week})" if latest_week else f"{THIS} 매출", eok(latest_sales))
cols[1].metric(f"{THIS} 수량", num(now_qty))
cols[2].metric(f"{THIS} 객단가", eok(now_avg))
cols[3].metric(f"{THIS} 수익", eok(now_profit), pct(now_rate))
if prev_week:
    wow_diff = latest_sales - prev_sales
    wow_sub = f"{PREV} 대비 {'+' if wow_diff >= 0 else '-'}{eok(abs(wow_diff))}"
    cols[4].metric(f"{latest_week} {WOW}", growth_pct(wow_rate), wow_sub, delta_color="off")
else:
    cols[4].metric(WOW, "-")

st.markdown(f"<div class='hint'>{THIS}({latest_week}) {len(fw):,}행 · 주문 {order_count:,}건 · 전체 로드 {len(f):,}행 · 기간 {f['날짜'].min().date()} ~ {f['날짜'].max().date()}</div>", unsafe_allow_html=True)

# -----------------------------
# 상단 개요: 최근 3개월 추이 / 이번주 브랜드 TOP10 / 이번주 공식·병행 비중
# -----------------------------
st.markdown("<div class='section-title'>개요</div>", unsafe_allow_html=True)
ov1, ov2 = st.columns([1, 1.5])

# (1) 최근 3주 매출 추이 (공식/병행 누적)
with ov1:
    wk_tp = f.groupby(["주차", "공식/병행"], as_index=False)["최종판매가"].sum()
    wk_tp["주차"] = pd.Categorical(wk_tp["주차"], categories=week_order, ordered=True)
    wk_tp = wk_tp.sort_values("주차")
    wk_tp["라벨"] = wk_tp["최종판매가"].apply(eok)
    fig_w = px.bar(
        wk_tp, x="주차", y="최종판매가", color="공식/병행", barmode="stack",
        text="라벨", title=f"{RECENT_TREND_TITLE} (공식/병행)",
        labels={"최종판매가": f"{INTERVAL_LABEL} 매출", "주차": PERIOD_AXIS},
        category_orders={"주차": week_order, "공식/병행": ["병행", "공식"]},
        color_discrete_map={"공식": "#2563eb", "병행": "#f59e0b"},
    )
    fig_w.update_traces(textposition="inside", textangle=0)
    fig_w.update_layout(xaxis_type="category", legend_title_text="공식/병행")
    wk_tot = f.groupby("주차")["최종판매가"].sum().reindex(week_order)
    fig_w.add_trace(go.Scatter(
        x=list(wk_tot.index), y=wk_tot.values,
        text=[eok(v) for v in wk_tot.values], mode="text",
        textposition="top center", textfont=dict(size=13, color="#0f172a"),
        showlegend=False, hoverinfo="skip",
    ))
    if len(wk_tot):
        ymax = float(wk_tot.max()); ymin = float(wk_tot.min())
        fig_w.update_yaxes(range=[min(0, ymin) * 1.1, ymax * 1.2 if ymax > 0 else ymax * 0.8])
    fig_w.update_yaxes(tickformat=",")
    st.plotly_chart(fig_w, use_container_width=True)

# (2) 이번주 브랜드 TOP10 (막대 안에 비중%만 크게)
with ov2:
    btop = aggregate(fw, ["브랜드"], metric_cols).head(10).copy()
    btop["라벨"] = btop["매출비중"].apply(lambda x: f"{x:.1f}%")
    fig_b = px.bar(
        btop, x="브랜드", y="최종판매가", text="라벨",
        title=f"{THIS} 브랜드 TOP 10 (매출 비중)", labels={"최종판매가": f"{THIS} 매출"},
    )
    fig_b.update_traces(
        textposition="inside", insidetextanchor="middle", textangle=0,
        textfont=dict(size=16, color="#ffffff"),
        marker_color="#2563eb",
        cliponaxis=False,
    )
    fig_b.update_layout(xaxis_type="category", margin=dict(t=54, b=10),
                        font=dict(size=14), title_font=dict(size=18),
                        uniformtext_minsize=11, uniformtext_mode="hide")
    fig_b.update_xaxes(categoryorder="total descending", tickangle=0, title_text="", tickfont=dict(size=12))
    fig_b.update_yaxes(title_text="", tickformat=",")
    st.plotly_chart(fig_b, use_container_width=True)

# -----------------------------
# 🎯 이번달 목표 달성률 (병행/공식별 쇼핑몰) — 주간/월간 무관, 이번달 기준
# -----------------------------
if not targets_df.empty:
    _cur_dt = df_all["날짜"].max()
    CUR_Y, CUR_M = int(_cur_dt.year), int(_cur_dt.month)
    month_df = df_all[(df_all["날짜"].dt.year == CUR_Y) & (df_all["날짜"].dt.month == CUR_M)].copy()
    month_df["_key"] = month_df["쇼핑몰"].map(_norm_mall)
    tgt_m = targets_df[targets_df["월"] == CUR_M]

    st.markdown(f"<div class='section-title'>🎯 {CUR_M}월 목표 달성률</div>", unsafe_allow_html=True)
    tot_target = float(tgt_m["목표"].sum())
    tot_actual = float(month_df["최종판매가"].sum())
    tot_rate = (tot_actual / tot_target * 100) if tot_target else None
    kc = st.columns(3)
    kc[0].metric(f"{CUR_M}월 실제 (업로드분)", eok(tot_actual))
    kc[1].metric(f"{CUR_M}월 목표", eok(tot_target))
    kc[2].metric("총 달성률", f"{tot_rate:.1f}%" if tot_rate is not None else "-")
    st.caption(f"업로드된 {CUR_M}월 데이터 기준 · 목표는 시트의 {CUR_M}월(연도 무시) · 목표 없는 쇼핑몰은 목표 0·달성률 '-' · 매출순 정렬")

    def _ach_table(block: str):
        bd = month_df[month_df["공식/병행"] == block]
        a = bd.groupby("_key")["최종판매가"].sum().to_dict()
        name_d = {}
        for _, r in bd.sort_values("최종판매가", ascending=False).iterrows():
            name_d.setdefault(r["_key"], r["쇼핑몰"])  # 데이터 쪽 대표 표시명
        t = tgt_m[tgt_m["공식병행"] == block]
        tgt_k = dict(zip(t["_key"], t["목표"]))
        name_t = dict(zip(t["_key"], t["쇼핑몰"]))
        keys = list(dict.fromkeys(list(t["_key"]) + list(a.keys())))  # 목표+실제 합집합(미매칭 포함)
        rows = []
        for k in keys:
            target = float(tgt_k.get(k, 0.0))
            actual = float(a.get(k, 0.0))
            name = name_t.get(k) or name_d.get(k) or k
            rate = (actual / target * 100) if target else None  # 목표 0 → 달성률 "-"
            rows.append({"쇼핑몰": name, "_t": target, "_a": actual, "_r": rate})
        tb = pd.DataFrame(rows)
        if len(tb):
            tb = tb.sort_values("_a", ascending=False)  # 매출순
        bt = float(tb["_t"].sum()) if len(tb) else 0.0
        ba = float(tb["_a"].sum()) if len(tb) else 0.0
        if not len(tb):
            return pd.DataFrame(columns=["쇼핑몰", "목표", "실제", "달성률"]), bt, ba
        body = pd.DataFrame({
            "쇼핑몰": tb["쇼핑몰"],
            "목표": pd.to_numeric(tb["_t"], errors="coerce").round(0),
            "실제": pd.to_numeric(tb["_a"], errors="coerce").round(0),
            "달성률": pd.to_numeric(tb["_r"], errors="coerce").round(1),
        })
        return body, bt, ba

    tcols = st.columns(2)
    for _col, _block in zip(tcols, ["병행", "공식"]):
        with _col:
            _abody, _abt, _aba = _ach_table(_block)
            _art = f"{_aba/_abt*100:.1f}%" if _abt else "-"
            st.markdown(f"<div class='section-title' style='font-size:15px;border:0'>{_block} · 합계 {eok(_abt)}→{eok(_aba)} ({_art})</div>", unsafe_allow_html=True)
            if PRINT_MODE:
                st.markdown(_df_to_html(_abody), unsafe_allow_html=True)
            else:
                st.dataframe(_abody, hide_index=True, use_container_width=True, column_config={
                    "목표": st.column_config.NumberColumn(format="%,.0f"),
                    "실제": st.column_config.NumberColumn(format="%,.0f"),
                    "달성률": st.column_config.NumberColumn(format="%.1f%%"),
                })
else:
    st.caption("🎯 목표 달성률을 보려면 '이미지.xlsx'를 앱 폴더에 두고 2번째 시트에 목표매출을 넣으세요. (현재 목표 시트를 못 읽었습니다)")


# -----------------------------
# 주차별 쇼핑몰 매출 (3주 비교)
# -----------------------------
st.markdown(f"<div class='section-title'>{PERIODLY} 쇼핑몰 매출 ({PERIOD_CMP})</div>", unsafe_allow_html=True)
wcols = st.columns(len(week_order)) if week_order else [st]
for col, wk in zip(wcols, week_order):
    with col:
        st.markdown(f"**{wk}**")
        wdf = f[f["주차"] == wk]
        wt = aggregate(wdf, ["쇼핑몰"], metric_cols).head(30).reset_index(drop=True)
        wt = wt[["쇼핑몰", "수량", "최종판매가", "수익률"]]
        wt.insert(0, "Rank", np.arange(1, len(wt) + 1))
        show_table(wt, height=620)

# -----------------------------
# 주차별 TOP 10 매출 상품 (3주 비교)
# -----------------------------
st.markdown(f"<div class='section-title'>{PERIODLY} TOP 10 매출 상품 ({PERIOD_CMP})</div>", unsafe_allow_html=True)
pcols = st.columns(len(week_order)) if week_order else [st]
for col, wk in zip(pcols, week_order):
    with col:
        st.markdown(f"**{wk}**")
        wdf = f[f["주차"] == wk]
        pt = aggregate(wdf, ["브랜드", "라인명"], metric_cols)
        st.markdown(product_cards_html(pt, n=10, img_px=84), unsafe_allow_html=True)

# =============================
# 이하 전부 '이번주(최신 주차)' 기준 — 주간보고용 (fw 는 위 KPI 블록에서 정의)
# =============================
st.markdown(
    f"<div class='section-title' style='margin-top:1.8rem;border-top:2px solid #e8eef5;padding-top:1rem;'>📅 {THIS} 상세 — {latest_week or ''}</div>",
    unsafe_allow_html=True,
)
st.caption(f"아래 섹션은 모두 {THIS} 데이터 기준입니다. ({PERIOD_CMP}는 위쪽 섹션 참고)")

# -----------------------------
# 공식/병행 (2분류) 분석
# -----------------------------
st.markdown(f"<div class='section-title'>공식/병행 분석 ({THIS})</div>", unsafe_allow_html=True)
type_table = aggregate(fw, ["공식/병행"], metric_cols).reset_index(drop=True)
type_table.insert(0, "Rank", np.arange(1, len(type_table) + 1))
tcol, _ = st.columns([1.4, 1])
with tcol:
    show_table(type_table, height=120)

# -----------------------------
# Brand deep analysis
# -----------------------------
st.markdown(f"<div class='section-title'>브랜드별 분석 ({THIS})</div>", unsafe_allow_html=True)


def brand_block(seg_df, title, topn):
    bt = aggregate(seg_df, ["브랜드"], metric_cols).head(topn).reset_index(drop=True)
    st.markdown(f"**{title}**")
    if bt.empty:
        st.caption("데이터 없음")
        return
    total = float(fw["최종판매가"].sum())  # 이번주 전체 매출 대비 비중
    bt["전체비중"] = bt["최종판매가"] / total * 100 if total else 0
    t = bt[["브랜드", "수량", "최종판매가", "전체비중", "수익률"]].copy()
    t.insert(0, "Rank", np.arange(1, len(t) + 1))
    show_table(t, height=80 + len(t) * 36)


bcol1, bcol2 = st.columns(2)
with bcol1:
    brand_block(fw[fw["공식/병행"] == "병행"], "병행 브랜드 TOP 10", 10)
with bcol2:
    brand_block(fw[fw["공식/병행"] == "공식"], "공식 브랜드 TOP 10", 10)

# -----------------------------
# Category deep analysis
# -----------------------------
st.markdown("<div class='section-title'>대분류별 분석</div>", unsafe_allow_html=True)
st.caption("현재 대분류 값: " + ", ".join(sorted(str(x) for x in fw["대분류"].unique())[:12]))
cat_table = aggregate(fw, ["대분류"], metric_cols)
cat_table = cat_table.head(50).reset_index(drop=True)
cat_table.insert(0, "Rank", np.arange(1, len(cat_table) + 1))

cg1, cg2 = st.columns([1.5, 1])
with cg1:
    st.plotly_chart(share_donut(cat_table, "대분류", "최종판매가", f"{THIS} 대분류 비중"), use_container_width=True)
with cg2:
    show_table(cat_table, height=520)

# -----------------------------
# 이번주 TOP 30 상품
# -----------------------------
st.markdown(f"<div class='section-title'>{THIS} TOP 30 상품</div>", unsafe_allow_html=True)
top30 = aggregate(fw, ["브랜드", "대분류", "라인명"], metric_cols).head(30).reset_index(drop=True)


def _top_mall_for(brand, line):
    """해당 상품(브랜드+라인명)이 가장 많이 팔린 쇼핑몰과 그 몰의 판매액 비중(%)."""
    sub = fw[(fw["브랜드"] == brand) & (fw["라인명"] == line)]
    if sub.empty or "쇼핑몰" not in sub.columns:
        return ("", float("nan"))
    g = sub.groupby("쇼핑몰")["최종판매가"].sum()
    g = g[g > 0]
    if g.empty:
        return ("", float("nan"))
    total = float(g.sum())
    return (str(g.idxmax()), (float(g.max()) / total * 100) if total else float("nan"))


_tm = [_top_mall_for(b, l) for b, l in zip(top30["브랜드"], top30["라인명"])]
top30["최다몰"] = [m for m, _ in _tm]
top30["최다몰비중"] = [p for _, p in _tm]
left = top30.iloc[0::2]   # 1, 3, 5, ... (왼쪽)
right = top30.iloc[1::2]  # 2, 4, 6, ... (오른쪽)
t30c1, t30c2 = st.columns(2)
with t30c1:
    st.markdown(product_cards_html(left, n=len(left), img_px=96, start=1, step=2), unsafe_allow_html=True)
with t30c2:
    st.markdown(product_cards_html(right, n=len(right), img_px=96, start=2, step=2), unsafe_allow_html=True)

# -----------------------------
# Performance tables, all descending by value
# -----------------------------
st.markdown("<div class='section-title'>성과 상세 분석</div>", unsafe_allow_html=True)
st.caption(f"‘{PREV}대비’ = {latest_week} vs {prev_week or '-'} 매출 증감({WOW})")

prev_wk_df = f[f["주차"] == prev_week] if prev_week else fw.iloc[0:0]
weekday_order = ["월", "화", "수", "목", "금", "토", "일"]
fw_wd = fw.copy()
fw_wd["요일"] = fw_wd["날짜"].dt.weekday.map({0: "월", 1: "화", 2: "수", 3: "목", 4: "금", 5: "토", 6: "일"})

tabs = st.tabs(["쇼핑몰", "브랜드", "대분류", "모델", "요일"])
specs = [("쇼핑몰", 50), ("브랜드", 50), ("대분류", 50), ("모델명", 100), ("요일", 7)]
for tab, (group_col, topn) in zip(tabs, specs):
    with tab:
        src = fw_wd if group_col == "요일" else fw
        table = aggregate(src, [group_col], metric_cols).reset_index(drop=True)
        if group_col not in ("요일",):  # 요일 탭은 WoW 의미 없음
            prev_g = prev_wk_df.groupby(group_col)["최종판매가"].sum()
            table[f"{PREV}매출"] = table[group_col].map(prev_g).fillna(0)
            _wow = np.where(table[f"{PREV}매출"] != 0,
                            (table["최종판매가"] - table[f"{PREV}매출"]) / table[f"{PREV}매출"].abs() * 100,
                            np.nan)
            table[f"{PREV}대비"] = pd.to_numeric(pd.Series(_wow), errors="coerce").replace([np.inf, -np.inf], np.nan)
            front = [group_col, "최종판매가", f"{PREV}매출", f"{PREV}대비"]
            table = table[front + [c for c in table.columns if c not in front]]
        if group_col == "요일":  # 월~일 순서로 정렬
            table["__o"] = table["요일"].map({d: i for i, d in enumerate(weekday_order)})
            table = table.sort_values("__o").drop(columns="__o").reset_index(drop=True)
        table = table.head(topn).reset_index(drop=True)
        if group_col == "쇼핑몰":
            table = rank_table(table, "쇼핑몰")
        else:
            table.insert(0, "Rank", np.arange(1, len(table) + 1))
        show_table(table, height=560)

# -----------------------------
# Summary
# -----------------------------
st.markdown("<div class='section-title'>자동 요약</div>", unsafe_allow_html=True)
mall_top = aggregate(fw, ["쇼핑몰"], metric_cols).head(1)
brand_top = aggregate(fw, ["브랜드"], metric_cols).head(1)
cat_top = aggregate(fw, ["대분류"], metric_cols).head(1)

week_totals = f.groupby("주차")["최종판매가"].sum()  # 3주 추이는 비교 맥락으로 유지
week_totals = week_totals.reindex([w for w in week_order if w in week_totals.index])

now_qty = fw["수량"].sum()
now_avg = latest_sales / now_qty if now_qty else 0

summary_lines = []
summary_lines.append(f"- {THIS}({latest_week}) 매출은 **{eok(latest_sales)}**, 수량 **{num(now_qty)}개**, 객단가 **{eok(now_avg)}**입니다.")
if len(week_totals) > 0:
    trend_str = " · ".join(f"{w} {eok(v)}" for w, v in week_totals.items())
    summary_lines.append(f"- {PERIODLY} 매출 추이({TREND_N}): {trend_str}")
if latest_week and prev_week:
    summary_lines.append(f"- {PREV} 대비 **{growth_pct(wow_rate)}** ({WOW})입니다.")
if not mall_top.empty:
    summary_lines.append(f"- 쇼핑몰 1위는 **{mall_top.iloc[0]['쇼핑몰']}**로 {THIS} **{eok(mall_top.iloc[0]['최종판매가'])}**, 비중 **{pct(mall_top.iloc[0].get('매출비중', np.nan))}**입니다.")
if not brand_top.empty:
    summary_lines.append(f"- 브랜드 1위는 **{brand_top.iloc[0]['브랜드']}**로 {THIS} **{eok(brand_top.iloc[0]['최종판매가'])}**입니다.")
if not cat_top.empty:
    summary_lines.append(f"- 대분류 1위는 **{cat_top.iloc[0]['대분류']}**로 {THIS} **{eok(cat_top.iloc[0]['최종판매가'])}**입니다.")
if not type_table.empty:
    parts = [f"{r['공식/병행']} {eok(r['최종판매가'])}({pct(r.get('매출비중', np.nan))})" for _, r in type_table.iterrows()]
    summary_lines.append("- 공식/병행 구성: " + " · ".join(parts) + "입니다.")

st.markdown("\n".join(summary_lines))

# -----------------------------
# 재고 현황 (재고 파일이 있을 때만)
# -----------------------------
if "stock_df" in dir() and isinstance(stock_df, pd.DataFrame) and not stock_df.empty:
    stock_df = stock_df[~stock_df["브랜드"].astype(str).str.strip().isin(_GIFT_BRANDS)].copy()  # 쇼핑백·사은품 제외(매출과 동일)
if "stock_df" in dir() and isinstance(stock_df, pd.DataFrame) and not stock_df.empty:
    st.markdown("<div class='section-title'>📦 재고 현황</div>", unsafe_allow_html=True)
    _sq = pd.to_numeric(stock_df["수량"], errors="coerce").fillna(0)
    _sc = pd.to_numeric(stock_df["총원가"], errors="coerce").fillna(0)
    _is_b = stock_df["공식/병행"].eq("병행") if "공식/병행" in stock_df.columns else pd.Series(True, index=stock_df.index)
    s1, s2, s3, s4 = st.columns(4)
    s1.metric("총 재고수량", f"{int(_sq.sum()):,}개")
    s2.metric("총 재고원가", eok(_sc.sum()))
    s3.metric("병행 원가", eok(_sc[_is_b].sum()))
    s4.metric("공식 원가", eok(_sc[~_is_b].sum()))

    def _stock_brand_top(d, n=30):
        gg = d.copy()
        gg["수량"] = pd.to_numeric(gg["수량"], errors="coerce").fillna(0)
        gg["총원가"] = pd.to_numeric(gg["총원가"], errors="coerce").fillna(0)
        t = (gg.groupby("브랜드", dropna=False)
               .agg(수량=("수량", "sum"), 총원가=("총원가", "sum"), 라인수=("라인명", "nunique"))
               .reset_index()
               .sort_values("총원가", ascending=False)
               .head(n)
               .reset_index(drop=True))
        if t.empty:
            return t
        t.insert(0, "순위", np.arange(1, len(t) + 1))
        return t[["순위", "브랜드", "총원가", "수량", "라인수"]]

    byo, gong = stock_df[_is_b], stock_df[~_is_b]
    cby, cgo = st.columns(2)
    with cby:
        st.markdown(f"**병행 — 총원가 TOP 30** · 합계 {eok(_sc[_is_b].sum())}")
        _tb = _stock_brand_top(byo)
        if PRINT_MODE:
            st.markdown(_df_to_html(_tb), unsafe_allow_html=True)
        else:
            st.dataframe(_tb, hide_index=True, use_container_width=True, column_config=_table_config(_tb))
    with cgo:
        st.markdown(f"**공식 — 총원가 TOP 30** · 합계 {eok(_sc[~_is_b].sum())}")
        _tg = _stock_brand_top(gong)
        if PRINT_MODE:
            st.markdown(_df_to_html(_tg), unsafe_allow_html=True)
        else:
            st.dataframe(_tg, hide_index=True, use_container_width=True, column_config=_table_config(_tg))
    st.caption(f"총 {len(stock_df):,}개 모델 · {stock_df['브랜드'].nunique():,}개 브랜드 · "
               f"병행 {int(_is_b.sum()):,} / 공식 {int((~_is_b).sum()):,} (쇼핑백·사은품 제외)")

# -----------------------------
# Download
# -----------------------------
st.markdown("<div class='section-title'>데이터 다운로드</div>", unsafe_allow_html=True)
csv = f.to_csv(index=False).encode("utf-8-sig")
st.download_button(
    "필터 적용 데이터 CSV 다운로드",
    data=csv,
    file_name="filtered_mecca_dashboard_data.csv",
    mime="text/csv",
)

with st.expander("컬럼 확인"):
    st.write(list(df.columns))
